#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""线体问题点提取 v3 — 智能解析 + 泰→中翻译双份

结构:
  文件名含日期: 8-20日制程问题点.xlsx (每天一个文件)
  每个工作表 = 一个车间 (RPO1/PRO2/...)
  每个单元格 = 一个系列的完整文本 (含多个时段区块):
    "FINAL C line\nTime 08:00-09:00\nTarget:444\nActual:359\nProblems-85\n-งานไม่มี stock...\nTime 09:00-10:00..."

解析逻辑:
  1. 按 "Time" 切分时段区块
  2. 每个区块内提取 Target/Actual/Problems(影响)
  3. 问题点描述 = 区块内 "ปัญหา|Problems|Problem" 之后的内容 (去掉线体名/时间/数字标签行)
  4. 泰语原文存 problem_th, 翻译成中文存 problem_zh (Google 免费接口)
  5. 智能过滤: 忽略 "ไม่พบปัญหาเครื่องจักร/ปกติ" 等无实质问题 (impact=0 或纯无问题描述)

提取结果写入 js/live-data.js 的 problems[]:
  {date, ws, series, line, time, problem_th, problem_zh, impact, plan, actual}
"""
import os, sys, re, glob, datetime, urllib.request, urllib.parse, json, time
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from live_data_lib import load_config, log, read_live_data, write_live_data, to_wsl_path

THAI_DIGITS = str.maketrans("๐๑๒๓๔๕๖๗๘๙", "0123456789")

# ── 专业知识库: 压缩机工厂术语对照 (泰→中) ──
# 来源: 效率管控文件夹 5月LOSS明细 中文.xlsx 高频词 + 精益生产专业术语
TERM_GLOSSARY = """压缩机/空调工厂专业术语 (翻译时优先使用):
ปะเต็ง = 碰焊/铆接   เชื่อม = 焊接   ตัด = 切割   ขัน = 拧紧
ลวด = 焊丝/线材   สายไฟ = 电线   ท่อ = 管路   ราง = 轨道
มอเตอร์ = 电机   เซ็นเซอร์ = 传感器   เกจ = 量规/检具   เมกะโอห์ม = 兆欧表(绝缘电阻)
ระบายไม่ทัน = 来不及排出   ทำความเย็นไม่ทัน = 制冷来不及   รั่ว = 泄漏
คัดแยก = 分选/挑选   ปน = 混入   แตก = 破裂   ค้าง = 卡滞   ติด = 卡住/粘住
หลวม = 松动   ไม่นิ่ง = 不稳定   ปรับเซ็ต = 调整设定   ไม่อ่านค่า = 读不到数值
ไม่ชำนาญ = 不熟练   อบรม = 培训   ขาดคน = 缺人   หยุด = 停机/停止
ไม่มี stock = 无库存   สต็อก = 库存   ไฟดับ = 断电   ไฟตก = 电压波动
บ่อย = 频繁   ซ้ำ = 重复   น้อย = 不足   งาน = 作业/工件
ประกอบ = 组装   เปลี่ยน = 更换   จัด = 整理   ความต่อเนื่อง = 连续性
หยุดเครื่อง = 设备停机   เครื่องเสีย = 设备故障   พัง = 损坏   ซ่อม = 维修
ตัวดัน = 推料机构   ลิฟท์ = 升降机   พาเลท = 托盘   คลีบ/คีบ = 夹爪/夹具
ตรวจสอบ = 检查   ตั้งค่า = 设定   ผลิต = 生产   แก้ไข = 修复   สกรู = 螺丝
โบลต์/น็อต = 螺栓/螺母   เพลา = 轴   ตลับลูกปืน = 轴承   โซ่ = 链条
สายพาน = 皮带   กระบอกสูบ = 气缸   น้ำมัน = 油   อากาศ = 气
ฉีด = 注射/喷涂   ทาสี = 涂漆   เช็ด = 擦拭   ล้าง = 清洗
"""

def _deepseek_key():
    """从 OpenClaw 配置读取 DeepSeek API key"""
    try:
        cfg_path = os.path.expanduser("~/.openclaw/openclaw.json")
        with open(cfg_path, encoding="utf-8") as f:
            cfg = json.load(f)
        return cfg["models"]["providers"]["deepseek"]["apiKey"]
    except Exception:
        return ""

_DS_KEY = _deepseek_key()

def translate_th_ai(text):
    """泰语→中文 (DeepSeek AI 翻译 + 工厂术语库, 失败回退 Google 机翻)"""
    if not text or not re.search(r"[\u0E00-\u0E7F]", text):
        return text
    if _DS_KEY:
        try:
            body = json.dumps({
                "model": "deepseek-chat",
                "messages": [
                    {"role": "system", "content": (
                        "你是泰国压缩机工厂的精益生产翻译专家。把泰语生产问题描述翻译成简洁专业的中文。\n"
                        "规则：1) 优先使用下方术语表中的专业术语；2) 去除客套与冗余，保留时间/数量/线体等关键信息；"
                        "3) 语气像现场班长上报，直接说问题+影响；4) 若原文含中/英文片段，保留原意并入译文。\n"
                        "术语表：\n" + TERM_GLOSSARY
                    )},
                    {"role": "user", "content": text[:800]}
                ],
                "temperature": 0.2,
                "max_tokens": 300
            }).encode("utf-8")
            req = urllib.request.Request("https://api.deepseek.com/chat/completions", data=body,
                headers={"Content-Type": "application/json", "Authorization": "Bearer " + _DS_KEY})
            with urllib.request.urlopen(req, timeout=30) as r:
                d = json.loads(r.read().decode("utf-8"))
            out = d["choices"][0]["message"]["content"].strip()
            # 去掉 AI 可能的引号包裹
            out = out.strip('"\'')
            if out:
                return out
        except Exception as e:
            log(f"AI翻译失败({e}), 回退Google")
    return translate_th(text)

def translate_th(text):
    """泰语→中文 (Google 免费接口, 失败返回原文) — 保留作回退"""
    if not text or not re.search(r"[\u0E00-\u0E7F]", text):
        return text
    try:
        url = "https://translate.googleapis.com/translate_a/single?client=gtx&sl=th&tl=zh-CN&dt=t&q=" + urllib.parse.quote(text[:900])
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=12) as r:
            data = json.loads(r.read().decode("utf-8"))
        out = "".join(seg[0] for seg in data[0] if seg[0])
    except Exception as e:
        log(f"翻译失败: {e}")
        return text
    # ── 后处理词典: 修正常见误译 (泰国工厂术语) ──
    fixes = [
        ("固定", "已修复"),
        ("修理", "已修复"),
        ("普通的", "正常"),
        ("普通", "正常"),
        ("卷板机", "线圈机"),
        ("电源不停", "断电停机"),
        ("没有电源", "断电"),
        ("没有库存", "无库存"),
        ("第一期", "初期"),
        ("造型", "调整"),
        ("刷动", "滑丝"),
        ("旋出", "滑丝"),
    ]
    for a, b in fixes:
        out = out.replace(a, b)
    return out

def today_str():
    return datetime.date.today().strftime("%Y-%m-%d")

def find_file(cfg):
    path = to_wsl_path(cfg.get("problems_excel_path") or "")
    if not path:
        return None, "未配置 problems_excel_path"
    if os.path.isfile(path):
        return path, None
    if not os.path.isdir(path):
        return None, f"路径不存在: {path}"
    pats = cfg.get("problems_excel_pattern") or ["问题点"]
    today = datetime.date.today()
    day_pats = [today.strftime("%-m-%-d"), today.strftime("%-m.%d"), today.strftime("%m-%d")]
    cands = []
    for f in os.listdir(path):
        if not f.lower().endswith((".xlsx", ".xlsm")):
            continue
        if not any(p in f for p in pats):
            continue
        if f.startswith("~$"):
            continue
        cands.append(f)
    if not cands:
        return None, f"目录 {path} 里没有匹配 {pats} 的Excel"
    for f in cands:
        if any(dp in f for dp in day_pats):
            return os.path.join(path, f), None
    latest = max(cands, key=lambda f: os.path.getmtime(os.path.join(path, f)))
    return os.path.join(path, latest), f"未找到今日文件, 使用最新: {latest}"

# ── 无实质问题 的泰语短语 (过滤掉, 不算问题点) ──
NO_PROBLEM_PATTERNS = [
    "ไม่พบปัญหา", "ไม่พบปญหา", "ปกติ", "ไม่มีปัญหา", "ไม่มีปญหา",
    "ทุกอย่างปกติ", "ทำงานปกติ", "โอเค", "ok",
]

def is_real_problem(desc):
    """判断是否实质问题 (不是'无问题/正常')"""
    d = (desc or "").strip().lower()
    if not d:
        return False
    for p in NO_PROBLEM_PATTERNS:
        if p in d:
            return False
    # 纯数字/符号/空壳
    d2 = re.sub(r"[\d\s\-+:,.;()\[\]%฿]", "", d)
    if not d2:
        return False
    return True

def split_blocks(text):
    """按 Time 切分时段区块, 返回 [(time_range, block_text), ...]"""
    t = (text or "").translate(THAI_DIGITS).replace("\r", "")
    # 用 Time/เวลา 切分 (时间标签后可能无冒号: "Time 08:00-09:00")
    parts = re.split(r"\n?\s*(?=Time\s*[:：]?\s*[\d]{1,2}[:：]|เวลา\s*[:：]?\s*[\d]|ช่วงเวลา\s*[:：]?\s*[\d])", t)
    blocks = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 找到时间标签 (可能在系列名之后, 用 search 不用 match)
        m = re.search(r"(?:Time|เวลา|ช่วงเวลา)\s*[:：]?\s*([\d]{1,2}[:：][\d]{2}\s*[-–~]\s*[\d]{1,2}[:：][\d]{2})", part)
        if m:
            time_range = m.group(1)
            # 去掉时间标签及其前面的系列名, 只留标签后的内容
            body = part[m.end():].strip()
            # 去掉 body 开头残留的系列名 (如 "F-series ")
            body = re.sub(r"^[A-Za-z][A-Za-z0-9\- ]*?\n", "", body)
            blocks.append((time_range, body))
        else:
            # 无时间标签的尾巴 (合并到上一个)
            if blocks:
                blocks[-1] = (blocks[-1][0], blocks[-1][1] + "\n" + part)
    return blocks

def extract_block_info(block):
    """从区块文本提取 Target/Actual/Problems + 问题描述"""
    b = block.translate(THAI_DIGITS)
    target = actual = impact = None
    # Target / Actual / Problems
    mt = re.search(r"[Tt]arget\s*[:：]?\s*([\d,]+)", b)
    ma = re.search(r"[Aa]ctual\s*[:：]?\s*([\d,]+)", b)
    if mt: target = int(mt.group(1).replace(",", ""))
    if ma: actual = int(ma.group(1).replace(",", ""))
    if target is not None and actual is not None:
        impact = actual - target
    # Problems: -85 / Diff : 1
    if impact is None:
        mp = re.search(r"(?:Problems|Problem|Diff)\s*[:：]?\s*([+-]?[\d,]+)", b)
        if mp:
            impact = int(mp.group(1).replace(",", ""))
    # 问题描述: "ปัญหา|Problems|Problem|เหตุผล" 之后的内容
    desc = ""
    mdesc = re.search(r"(?:ปัญหา|ปญหา|Problems|Problem|เหตุผล|สาเหตุ)\s*[:：]?\s*(.*)", b, re.S)
    if mdesc:
        desc = mdesc.group(1).strip()
    # 去掉描述里的 -85 等影响数字行首 / 纯数字行
    lines = []
    for ln in desc.splitlines():
        l2 = ln.strip().lstrip("-–•* ")
        if not l2:
            continue
        if re.fullmatch(r"[+-]?\d[\d,]*", l2):
            continue  # 纯数字行 (如 "-85")
        lines.append(l2)
    desc = " ".join(lines).strip()
    # 清理: 去掉重复的线体名头
    desc = re.sub(r"^(FINAL\s+[A-Z]\s*LINE|Final\s+[A-Z]|S-series line.*?)\s*[:：]?\s*", "", desc, flags=re.I)
    # 去掉尾部影响数字 "-85" / "(-20)" / "-33" (行尾残留)
    desc = re.sub(r"\s*\(?[-−]\d{1,4}\)?\s*$", "", desc).strip()
    return target, actual, impact, desc

def extract_ws(ws, date_str, translate=True):
    """提取单个工作表(车间)的问题点 — v3: 按单元格全文智能解析"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    entries = []
    # 找表头行: 含「问题点」
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(v) if v is not None else "" for v in row]
        if any("问题点" in v or "problem" in v.lower() or "ปัญหา" in v for v in vals):
            header_row = i
            break
    if header_row is None:
        # 兜底: 第1行就是系列名, 第2行是问题点
        if len(rows) >= 2 and any(str(v) for v in rows[1]):
            header_row = 1
        else:
            return []
    hdr = rows[header_row]
    # 系列名: 表头上一行
    series_row = rows[header_row - 1] if header_row > 0 else None
    for ci, v in enumerate(hdr):
        if v is None:
            continue
        vs = str(v).strip()
        if vs != "问题点" and "problem" not in vs.lower() and "ปัญหา" not in vs:
            continue
        # 该列所有数据行 (通常就是下一行, 也可能多行)
        series = ""
        if series_row is not None and ci < len(series_row) and series_row[ci] is not None:
            series = str(series_row[ci]).strip()
        for ri in range(header_row + 1, len(rows)):
            row = rows[ri]
            if ci >= len(row) or row[ci] is None:
                continue
            cell = str(row[ci]).strip()
            if not cell:
                continue
            blocks = split_blocks(cell)
            for time_range, block in blocks:
                tg, ac, imp, desc = extract_block_info(block)
                # 过滤无实质问题 (impact 0 或 无问题描述) — 但要保留有影响的实际问题
                if not desc:
                    continue
                if not is_real_problem(desc):
                    continue
                desc_th = desc
                desc_zh = translate_th_ai(desc) if translate else desc
                line_name = (ws.title + "·" + series) if series else ws.title
                entries.append({
                    "date": date_str,
                    "ws": ws.title,
                    "series": series,
                    "line": line_name,
                    "time": time_range,
                    "problem_th": desc_th,
                    "problem_zh": desc_zh,
                    "plan": tg,
                    "actual": ac,
                    "impact": imp,
                })
    return entries

def ai_aggregate_top3(entries):
    """AI 语义判断 + 精准分组 TOP3.

    2026-08-21 13:34 用户最终要求:
    - 不是逐字相同才合并! 同一问题用不同表述(换说法) AI 也要认出来是同一个问题.
    - 但必须精准严谨: 不同的独立问题绝不合并, 不确定就不要合并.
    - AI 只负责「判断哪些条目是同一个问题」的分组, 组名/描述一律用原始精准翻译(problem_zh),
      绝不让 AI 自创概括名字 (防止笼统化).
    - 输出按累计影响绝对值降序 TOP3.
    失败回退: 返回 None → 前端退回全量 problems.
    """
    if not entries:
        return None
    if len(entries) < 2:
        return None
    # 送 AI 的条目: 索引 + 线体 + 时段 + 影响 + 泰语原文 + 中文翻译
    items = []
    for i, p in enumerate(entries):
        items.append({
            "i": i,
            "line": p.get("line", ""),
            "time": p.get("time", ""),
            "impact": int(p.get("impact") or 0),
            "th": (p.get("problem_th") or "").strip()[:150],
            "zh": (p.get("problem_zh") or "").strip()[:150],
        })
    try:
        body = json.dumps({
            "model": "deepseek-chat",
            "messages": [
                {"role": "system", "content": (
                    "你是泰国压缩机工厂的生产管理分析师。下面是今天各时段上报的问题点条目(每条含: 索引i、线体、时段、影响台数、泰语原文th、中文翻译zh)。\n"
                    "任务: 把「本质上是同一个问题」的条目分组。\n"
                    "规则:\n"
                    "1) 同一个问题允许用不同话表述——例如\"等待换型\"和\"换型中停产\"、\"按钮坏了\"和\"按钮故障\"、同一型号换型在不同时段反复出现, 都算同一个问题。你要用语义理解判断。\n"
                    "2) 不同的问题绝不能合并——即使话题相近(如\"缺料\"和\"缺人\"是两个问题), 不确定是否同一问题时不要合并。宁可分组更细, 不可错误合并。\n"
                    "3) 不同线体(line字段)的条目默认不能合并——除非能100%确定是同一台共用设备/同一个问题跨线体影响(例如同一台m线圈设备故障同时影响多条线), 不确定就分开。\n"
                    "4) 每条条目只能归入一组, 同一组的条目必须是同一个根本问题。\n"
                    "5) 只输出 JSON 数组, 每组: {\"items\":[该组条目索引数组]}。不要输出组名、不要解释、不要 markdown 代码块。\n"
                    "例: [{\"items\":[0,3]},{\"items\":[1,4,5]},{\"items\":[2]}]"
                )},
                {"role": "user", "content": json.dumps(items, ensure_ascii=False)[:8000]}
            ],
            "temperature": 0.1,
            "max_tokens": 800
        }).encode("utf-8")
        req = urllib.request.Request("https://api.deepseek.com/chat/completions", data=body,
            headers={"Content-Type": "application/json", "Authorization": "Bearer " + _DS_KEY})
        with urllib.request.urlopen(req, timeout=60) as r:
            d = json.loads(r.read().decode("utf-8"))
        out = d["choices"][0]["message"]["content"].strip()
        out = re.sub(r"^```(?:json)?\s*|\s*```$", "", out, flags=re.S).strip()
        arr = json.loads(out)
        if not isinstance(arr, list):
            return None
        # 组装分组: 组名 = 组内累计影响最大那条的原始精准翻译, 不用 AI 自创名
        top3 = []
        for g in arr:
            if not isinstance(g, dict):
                continue
            idxs = [x for x in (g.get("items") or []) if isinstance(x, int) and 0 <= x < len(entries)]
            idxs = list(dict.fromkeys(idxs))  # 去重保序
            if not idxs:
                continue
            group_entries = [entries[i] for i in idxs]
            total_impact = sum(int(p.get("impact") or 0) for p in group_entries)
            # 代表条目 = 影响绝对值最大的一条
            rep = max(group_entries, key=lambda p: abs(int(p.get("impact") or 0)))
            name = (rep.get("problem_zh") or rep.get("problem_th") or "").strip()[:120]
            # 2026-08-21 13:57: 英文界面显示泰语原文 (用户此前要求, 重构时遗漏) — 同时带泰语名
            name_th = (rep.get("problem_th") or "").strip()[:120]
            lines = []
            times = []
            for p in group_entries:
                if p.get("line") and p["line"] not in lines:
                    lines.append(p["line"])
                if p.get("time") and p["time"] not in times:
                    times.append(p["time"])
            top3.append({
                "name": name,
                "name_th": name_th,
                "lines": lines,
                "times": times,
                "count": len(idxs),
                "total_impact": total_impact,
            })
        top3.sort(key=lambda x: abs(x["total_impact"]), reverse=True)
        top3 = top3[:3]
        if top3:
            log(f"AI语义分组TOP3完成: " + " | ".join(f"{x['name'][:24]}({x['total_impact']})x{x['count']}" for x in top3))
            return top3
    except Exception as e:
        log(f"AI语义分组失败({e}), 保留全量problems")
    return None


def main():
    cfg = load_config()
    path, warn = find_file(cfg)
    if not path:
        log(f"⚠️ {warn}")
        return 0
    if warn:
        log(f"⚠️ {warn}")
        # ── 2026-08-25: 兜底 — 找不到今日文件时自动补建当天模板 ──
        #    背景: 8:05 的 create_daily_problem_file cron 依赖机器开机时间,
        #    电脑开机晚几秒就会错过 8:05 → 一整天都停在旧文件。
        #    这里每10分钟检查一次, 只要发现当天文件缺失就自动创建, 最迟 8:10 补上。
        try:
            import create_daily_problem_file as _cdp
            full, created = _cdp.ensure_daily_file(cfg)
            if full:
                path, warn = find_file(cfg)
                log(f"🛟 自动补建当天文件: {os.path.basename(path)}" + (" (新建)" if created else " (已存在)"))
        except Exception as e:
            log(f"⚠️ 自动补建当天文件失败: {e}")
    # ── 2026-08-24: 变化检测 — 文件签名(路径+mtime+size)未变则跳过, 避免空转调 AI
    #    现场保存Excel后 mtime 必变 → 下次cron立即重新提取; 未保存则静默跳过
    state_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".extract_state.json")
    sig = {"file": path, "mtime": os.path.getmtime(path), "size": os.path.getsize(path)}
    state = {}
    try:
        with open(state_file, encoding="utf-8") as f:
            state = json.load(f)
    except Exception:
        pass
    if "--force" not in sys.argv and state.get("file") == sig["file"] and state.get("mtime") == sig["mtime"] and state.get("size") == sig["size"]:
        log(f"文件未变化, 跳过提取 ({os.path.basename(path)})")
        return 0
    date_str = today_str()
    m = re.search(r"(\d{1,2})[-.](\d{1,2})", os.path.basename(path))
    if m:
        try:
            d = datetime.date(datetime.date.today().year, int(m.group(1)), int(m.group(2)))
            date_str = d.strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log(f"打开Excel失败 {path}: {e}")
        return 1
    entries = []
    for name in wb.sheetnames:
        entries.extend(extract_ws(wb[name], date_str, translate=True))
    wb.close()
    data = read_live_data()
    # ── 2026-08-21: 只保留当天数据 (新的一天从空开始, 旧日期条目不再混入 Top3) ──
    #    历史数据由 save_daily_history.py 每日17:00归档到 history/*.json + __HISTORY__
    problems = []
    problems.extend(entries)
    data["problems"] = problems
    # ── 2026-08-21 13:30: AI 语义聚合主要问题点 TOP3 (同一问题跨时段合并累计影响) ──
    #    用户要求: 不是某一时段的最大值, 而是主要问题点; 前端优先渲染 problems_top, 无则退回全量
    data["problems_top"] = ai_aggregate_top3(entries)
    write_live_data(data)
    log(f"v5提取完成: {os.path.basename(path)} → 今日{date_str} 全量{len(entries)}条 (累计{len(problems)}条" + (f", AI TOP3={len(data['problems_top'])}条" if data.get("problems_top") else "") + ")")
    # 提取成功 → 更新签名状态 (失败则保留旧状态, 下次cron自动重试)
    try:
        with open(state_file, "w", encoding="utf-8") as f:
            json.dump(sig, f, ensure_ascii=False)
    except Exception as e:
        log(f"写入状态文件失败: {e}")
    return 0

if __name__ == "__main__":
    sys.exit(main())
